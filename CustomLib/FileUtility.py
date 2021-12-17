import openpyxl
import xlrd
from openpyxl import workbook
from openpyxl.reader.excel import load_workbook
import csv
import json
import pandas as pd

'''This utility works for reading data from different sources'''
'''Sources like csv, xlsx, xls'''


def read_xls_file(file_name, sheet_name):
    data_list = []
    with xlrd.open_workbook(file_name) as open_workbook:
        open_sheet = open_workbook.sheet_by_name(sheet_name)

        row_count = open_sheet.nrows
        col_count = open_sheet.ncols

        for curr_row in range(1, row_count):
            row_data = []
            for curr_col in range(0, col_count):
                row_data.append(open_sheet.cell_value(curr_row, curr_col))
            data_list.append(row_data)
        return data_list


def get_active_sheets(file_name):
    with workbook(file_name) as open_workbook:
        return open_workbook.active


def get_all_sheets(file_name):
    with workbook(file_name) as open_workbook:
        return open_workbook.sheetnames


def read_xlsx_file(file_name, sheet_name):
    # import pdb; pdb.set_trace()
    open_workbook = pd.read_excel(io=file_name, sheet_name=sheet_name)
    return {d['${USEREMAIL}']: d['${PASSWORD}'] for d in
            open_workbook.to_dict(orient='records')}  # dictonary comprehension


def read_xlsx_file2(file_name, sheet_name):
    data_list = []
    with open(file_name, 'rb') as xlsx_file:
        open_workbook = load_workbook(xlsx_file)
        open_sheet = open_workbook[sheet_name]

        row_count = open_sheet.max_row
        col_count = open_sheet.max_column

        for i in range(2, row_count + 1):
            row_data = []
            for j in range(1, col_count + 1):
                row_data.append(open_sheet.cell(i, j).value)
            data_list.append(row_data)
        return data_list


def read_columns_id(file_name, sheet_name, column_name):
    with workbook(file_name) as open_workbook:
        open_sheet = open_workbook[sheet_name]

        row_count = open_sheet.max_row
        col_count = open_sheet.max_column

        for i in range(1, col_count):
            if open_sheet.cell(1, i).value == column_name:
                return i
            break


def read_csv_file(file_name):
    data = []
    with open(file_name, "r") as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            data.append(row)
    return data


def read_json_file(file_name, node_name):
    with open(file_name) as open_file:
        data = json.load(open_file)

        for i in data[node_name]:
            pass
        return i